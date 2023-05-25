#!/usr/bin/env python3

import collections
import contextlib
import csv
import datetime
import io
import os.path
import random
import re
import sys
import zipfile
from abc import abstractmethod
from enum import Enum, auto
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, BinaryIO, Callable, ClassVar, Iterable, Iterator, NamedTuple, Optional, Type, TypeVar, Union

import toml
from openpyxl.reader.excel import load_workbook
# Note: openpyxl was chosen as it's the pandas dependency for loading xlsx documents.
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from gooey import Gooey, GooeyParser

ENCODED_DIGITS = 16
ENC_PATTERN = re.compile(r"enc-\d{16}")  # make sure this matches ENCODED_DIGITS
REPORT_PROGRESS = True

ConfigType = TypeVar('ConfigType', bound='BaseConfig')


def random_digits():
    return ''.join(str(random.randint(0, 9)) for _ in range(ENCODED_DIGITS))


class ZipPath(zipfile.Path):
    @property
    def stem(self) -> str:
        return Path(str(self)).stem

    @property
    def suffix(self) -> str:
        return Path(str(self)).suffix

    def iterdir(self) -> Iterator['ZipPath']:  # noqa (iterdir is not a word)
        for entry in super().iterdir():
            yield ZipPath.from_zip_path(entry)

    @classmethod
    def from_zip_path(cls, entry: zipfile.Path) -> 'ZipPath':
        return cls(
            entry.root.filename,  # noqa (root.filename is like a private interface)
            at=entry.at,  # noqa (at is like a private interface)
        )


FilePath = Union[Path, ZipPath]


class Operation(Enum):
    ENCODE = auto()
    DECODE = auto()


class QueueItem:
    def __init__(self, path: FilePath, config: 'BaseConfig', operation: Operation):
        self.path: FilePath = path
        self.config = config
        self.operation = operation

    def process(self, worker: 'Worker'):
        destination_buffer = self.config.make_destination_buffer()
        if self.operation == Operation.ENCODE:
            self.config.encode_file(self.path, worker, destination_buffer)
        else:
            self.config.decode_file(self.path, worker, destination_buffer)
        buffer_data = self.config.make_buffer_binary(destination_buffer)
        worker.save_output(self.output_name(), buffer_data)

        supporting_files = self.config.get_supporting_files(self.path)
        worker.save_supporting_files(supporting_files)

    def output_name(self) -> str:
        return self.path.name

    def __str__(self) -> str:
        return f'{self.path}'


class Worker:
    MAPPING_FILE_NAME = 'mapping.tsv'

    def __init__(self, output_directory: str, output_zipname: Optional[str] = None, should_save_mappings: bool = True):
        self.output_directory: Path = Path(output_directory)
        self.output_directory.mkdir(parents=True, exist_ok=True)

        self.encoded_mappings: dict[str, str] = {}
        self.encoded_values: set[str] = set()
        self.input_zipfiles: dict[Path, zipfile.ZipFile] = {}
        self.filesizes: list[int] = []
        self.queue: list[QueueItem] = []
        self.output_names: set[str] = set()
        output_zipname = output_zipname or 'output.zip'  # TODO: timestamped name by default?
        self.output_zipfile: zipfile.ZipFile = \
            zipfile.ZipFile(self.output_directory / output_zipname, mode="w", compression=zipfile.ZIP_DEFLATED)
        self.should_save_mappings = should_save_mappings
        self.processed_count: int = 0

    def unique_output_name(self, name: str):
        if name in self.output_names:
            suffix = 2
            while f'{name}.{suffix}' in self.output_names:
                suffix += 1
            name = f'{name}.{suffix}'
        self.output_names.add(name)
        return name

    def encoded_replace(self, match: re.Match):
        return self.encoded_mappings[match.group()]

    def save_supporting_files(self, in_files: list[FilePath]) -> None:
        # TODO: optimize
        #  This is far from being "ok". It assumes that the whole file is held in memory,
        #  it can change the output file name etc.
        for file_path in in_files:
            output_name = self.unique_output_name(file_path.name)
            with file_path.open(mode='rb') as f:  # noqa (mode is supported)
                self.output_zipfile.writestr(output_name, f.read())

    def save_output(self, path: str, content: bytes) -> None:
        output_name = self.unique_output_name(path)
        # writestr tries to encode string into UTF-8, so we pass bytes instead.
        assert isinstance(content, bytes)
        self.output_zipfile.writestr(output_name, content)

    def _list_files(self, paths: Iterable[str]) -> list[FilePath]:
        """
        Lists all files inside all provided paths, including inside of zip archives.

        We'll be using this later on to search for files in the same directories with e.g. header order list.
        """
        paths: collections.deque[FilePath] = collections.deque(Path(x) for x in paths)
        all_files = []
        while paths:
            path = paths.popleft()
            if not path.exists():
                print(f'{path} does not exist, skipping')
                continue

            if path.is_dir():
                paths.extend(path.iterdir())
                continue

            if path.suffix.lower().endswith('.zip'):
                paths.extend(ZipPath(path, '').iterdir())
                continue

            all_files.append(path)
        return all_files

    def find_files(self, paths: list, for_encode: bool) -> None:
        list_of_files = self._list_files(paths)
        print(f'Listed {len(list_of_files)} files.')
        for file_path in list_of_files:
            config = ConfigFactory.get_config(file_path.name)
            if config is None:
                continue

            self.queue.append(QueueItem(file_path, config, Operation.ENCODE if for_encode else Operation.DECODE))
            self.filesizes.append(file_path.stat().st_size)

    def encode_value(self, value: str) -> str:
        try:
            return self.encoded_mappings[value]
        except KeyError:
            while True:
                encoded = 'enc-' + random_digits()
                if encoded not in self.encoded_values:
                    self.encoded_values.add(encoded)
                    break
            self.encoded_mappings[value] = encoded
            return encoded

    def process_files(self):
        total = len(self.queue)
        total_file_size = sum(self.filesizes)
        processed_bytes = 0
        for queue_item, filesize in zip(self.queue, self.filesizes):
            self.processed_count += 1
            print(f'Processing file {queue_item} ({self.processed_count}/{total})')
            queue_item.process(self)
            processed_bytes += filesize
            if REPORT_PROGRESS:
                print(f'Progress {int((processed_bytes * 100) / total_file_size)}%')
        print(f'Successfully processed {self.processed_count} data files')

    def save_mappings(self):
        # TODO: write to temp and rename?
        with open(self.output_directory / self.MAPPING_FILE_NAME, mode="w", encoding='utf-8') as f:
            writer = csv.writer(f, dialect='excel-tab')
            writer.writerows(sorted(self.encoded_mappings.items()))

    def load_mappings(self, path):
        # No matter other encodings, mappings are always saved as `utf-8`.
        with open(path, mode="r", encoding='utf-8') as f:
            reader = csv.reader(f, dialect='excel-tab')
            self.encoded_mappings = dict((x[1], x[0]) for x in reader if len(x) == 2)
            self.encoded_values = set(self.encoded_mappings.values())

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.output_zipfile.close()
        for f in self.input_zipfiles.values():
            f.close()
        if self.should_save_mappings:
            self.save_mappings()


class BaseConfig:
    CONFIG_TYPE: ClassVar[str] = None
    BUFFER_TYPE: ClassVar[Type] = io.TextIOWrapper

    def __init__(self, file_mask: str, carrier: str, encoding: str = 'utf-8'):
        self.file_mask = file_mask
        self.carrier = carrier
        self.encoding = encoding

    def __str__(self) -> str:
        return f'{self.carrier} with mask {self.file_mask}'

    def matches(self, filename: str) -> bool:
        return re.match(self.file_mask, filename) is not None

    def get_supporting_files(self, in_file: FilePath) -> list[FilePath]:
        return []

    def make_destination_buffer(self) -> BUFFER_TYPE:
        return io.TextIOWrapper(buffer=io.BytesIO(), encoding=self.encoding)

    def make_buffer_binary(self, destination_buffer: BUFFER_TYPE) -> bytes:
        destination_buffer.seek(0)
        data = destination_buffer.read()
        return data.encode(self.encoding)

    @abstractmethod
    def encode_file(self, in_file: FilePath, worker: Worker, destination: BUFFER_TYPE) -> None:
        raise NotImplementedError

    @abstractmethod
    def decode_file(self, in_file: FilePath, worker: Worker, destination: BUFFER_TYPE) -> None:
        raise NotImplementedError

    @abstractmethod
    def get_description(self) -> dict[str, str]:
        raise NotImplementedError

    def make_description(self, message: str) -> dict[str, str]:
        return {
            'type': 'MessageDialog',
            'menuTitle': f'{self.carrier} - {self.file_mask}',
            'caption': f'Configuration for {self.carrier} - {self.file_mask}',
            'message': message,
        }


class ConfigFactory:
    REGISTERED: ClassVar[dict[str, Type[BaseConfig]]] = {}
    LOADED: ClassVar[list[BaseConfig]] = []
    COMMON_TAG: ClassVar[str] = 'common'
    DEFAULT_CONFIGURATION: ClassVar[Path] = Path(__file__).parent / Path('./config.toml')

    @classmethod
    def register(cls, config_class: Type[ConfigType]) -> Type[ConfigType]:
        cls.REGISTERED[config_class.CONFIG_TYPE] = config_class
        return config_class

    @classmethod
    def get_config(cls, filename: str) -> Optional[BaseConfig]:
        cls.load_configuration()
        return next((config for config in cls.LOADED if config.matches(filename)), None)

    @classmethod
    def get_config_descriptions(cls) -> Iterator[dict[str, str]]:
        cls.load_configuration()
        for config in cls.LOADED:
            yield config.get_description()

    @classmethod
    def load_configuration(cls) -> None:
        if len(cls.LOADED) > 0:
            return

        config_path = cls.DEFAULT_CONFIGURATION
        toml_data = toml.loads(config_path.read_text())

        for namespace, values_map in toml_data.items():
            common_values = values_map.get(cls.COMMON_TAG, {})

            for sub_namespace, parameters in values_map.items():
                if sub_namespace == cls.COMMON_TAG:
                    continue
                full_params: dict = parameters.copy()
                full_params.update(**common_values)

                config_class_name = full_params.pop('config_class')
                config_class = cls.REGISTERED[config_class_name]
                instance = config_class(**full_params)  # noqa
                cls.LOADED.append(instance)

        print(f'Loaded {len(cls.LOADED)} configuration options.')


class Condition(NamedTuple):
    replace_where: str
    if_column: str
    comparison_operator: str
    has_value: str

    def does_match(self, column_value: str) -> bool:
        operators = {
            '==': column_value == self.has_value,
            '!=': column_value != self.has_value,
        }
        try:
            return operators[self.comparison_operator]
        except KeyError as ex:
            raise ValueError(
                f'Unknown operator {self.comparison_operator}, use one of {list(operators.keys())}'
            ) from ex


class SingleReplacement(NamedTuple):
    value: str
    span_start: int
    span_end: int

    def apply(self, in_str: list[str]) -> None:
        in_str[self.span_start: self.span_end] = list(self.value)


class EncodeRegex:
    def __init__(self, expression: str):
        self.expression = expression

    def encode(self, value: str, encoder: Callable[[str], str]) -> str:
        result = re.search(self.expression, value)
        if result is None or len(result.groupdict()) == 0:
            return value

        # We need to replace elements in this string from the back, to ensure that
        # indices will always point to the right place in the output string.
        replacements = []
        for group_name, group_value in result.groupdict().items():
            span_start, span_end = result.span(group_name)
            replacement_value = encoder(group_value)
            replacements.append(SingleReplacement(replacement_value, span_start, span_end))

        # Replacing non-mutable string with a list of single characters that we'll work on.
        out_value = list(value)
        for replacement in sorted(replacements, key=lambda elem: elem.span_end, reverse=True):
            replacement.apply(out_value)

        return ''.join(out_value)


# Simplification for working with tables.
# Note: inheriting a NamedTuple is a pain.
class TableEncodeRegex(EncodeRegex):
    def __init__(self, expression: str, replace_where: str):
        super().__init__(expression)
        self.replace_where = replace_where


@ConfigFactory.register
class CSVConfig(BaseConfig):
    CONFIG_TYPE = 'csv-config'

    def __init__(
        self,
        clear_columns: Iterable[str],
        encode_columns: Iterable[str],
        encode_conditional: Optional[Iterable[list[str]]] = None,
        encode_regex: Optional[Iterable[list[str]]] = None,
        dialect: str = 'excel',
        delimiter: Optional[str] = None,
        num_headers: int = 1,
        skip_initial_lines: int = 0,
        external_header_file: Optional[str] = None,
        external_header_format: Optional[str] = None,
        **kwargs,
    ):
        super().__init__(**kwargs)
        self.dialect = dialect
        self.clear_columns = clear_columns
        self.encode_columns = encode_columns
        self.encode_conditional = [Condition(*entry) for entry in (encode_conditional or [])]
        self.encode_regex = [TableEncodeRegex(*entry) for entry in (encode_regex or [])]
        self.delimiter = delimiter
        self.num_headers = num_headers
        self.skip_initial_lines = skip_initial_lines
        self.external_header_file = external_header_file
        self.external_header_format = external_header_format

    def _process(
        self,
        in_file: FilePath,
        worker: Worker,
        destination: io.TextIOWrapper,
        mapper: Callable[[dict[str, str], Worker, dict[str, str]], dict[str, str]],
    ) -> None:
        with self.make_csv_reader_writer(in_file, destination) as (reader, writer):
            stripped_fieldnames = {key.strip(): key for key in reader.fieldnames}

            # In case of some operators, they can have multiple header rows at the start of the file.
            additional_headers = []
            while (len(additional_headers) + 1) < self.num_headers:
                additional_headers.append(next(reader))

            if self.external_header_file is None:
                writer.writeheader()
                # Write additional header lines back to the anonymized file.
                writer.writerows(additional_headers)

            for row in reader:
                mapped_row = mapper(row, worker, stripped_fieldnames)
                writer.writerow(mapped_row)

    def encode_file(self, in_file: FilePath, worker: Worker, destination: io.TextIOWrapper) -> None:
        self._process(in_file, worker, destination, self.mapper)

    def decode_file(self, in_file: FilePath, worker: Worker, destination: io.TextIOWrapper) -> None:
        self._process(in_file, worker, destination, self.de_mapper)

    def get_supporting_files(self, in_file: FilePath) -> list[FilePath]:
        if self.external_header_file is None:
            return []
        return [self._get_header_file_path(in_file)]

    def make_csv_config(self) -> dict[str, str]:
        config = {
            'dialect': self.dialect,
        }
        if self.delimiter:
            config['delimiter'] = self.delimiter
        return config

    def _get_header_file_path(self, in_file: FilePath) -> FilePath:
        return in_file.parent / self.external_header_file

    def _load_fieldnames(self, in_file: FilePath) -> Optional[list[str]]:
        if self.external_header_file is None:
            return None

        # It is assumed that external header will be placed relative to the original file with data.
        header_file_path = self._get_header_file_path(in_file)

        # Name of the handler is either carrier or specific name. We assume that in most cases
        # a single carrier will have a single format, but don't want to limit ourselves to that case.
        handler_fun = {
            'Rogers': self._load_rogers_fieldnames,
        }[self.external_header_format or self.carrier]
        return handler_fun(header_file_path)

    def _load_rogers_fieldnames(self, header_file: FilePath) -> list[str]:
        """
        Rogers header file contains:
        - a single line with header type name and period
        - an empty line
        - CSV with fields "field name", "max length", "type" in that order
        We're interested in "field name" only.
        """
        field_name_tag = 'field_name'
        out_list = []

        with header_file.open(mode='r', encoding=self.encoding) as f:  # noqa (encoding is supported)
            for _ in range(2):
                f.readline()

            reader = csv.DictReader(  # noqa (pathlib.Path / zipfile.Path open result is good enough)
                f,
                fieldnames=[field_name_tag, 'max length', 'type'],
                # This is specific to the header, normal files are separated by pipes.
                delimiter=',',
            )
            for line_dict in reader:
                out_list.append(line_dict[field_name_tag])

        return out_list

    @contextlib.contextmanager
    def make_csv_reader_writer(
        self,
        in_file: FilePath,
        destination: io.TextIOWrapper,
    ) -> tuple[csv.DictReader, csv.DictWriter]:
        fieldnames = self._load_fieldnames(in_file)
        with in_file.open(mode='r',
                          encoding=self.encoding) as source:  # noqa (all FilePath types support encoding on open)
            # We can have a header that doesn't provide any data. It's rewritten "as is".
            for _ in range(self.skip_initial_lines):
                line = source.readline()
                destination.writelines([line])  # noqa

            config = self.make_csv_config()
            reader = csv.DictReader(f=source, fieldnames=fieldnames, **config)  # noqa
            writer = csv.DictWriter(f=destination, fieldnames=fieldnames or reader.fieldnames, **config)
            yield reader, writer

    def mapper(
        self,
        in_data: dict[str, str],
        worker: Worker,
        fieldnames_mapping: dict[str, str],
    ) -> dict[str, str]:
        # It is possible that each key here requires striping.
        mapped_data = in_data.copy()
        encode = worker.encode_value

        for key in self.clear_columns:
            mapped_data[fieldnames_mapping[key]] = ''

        for key in self.encode_columns:
            mapped_data[fieldnames_mapping[key]] = encode(mapped_data.get(fieldnames_mapping[key]) or '')

        for condition in self.encode_conditional:
            column_value = mapped_data[fieldnames_mapping[condition.if_column]].strip()
            if condition.does_match(column_value):
                mapped_data[fieldnames_mapping[condition.replace_where]] = \
                    encode(mapped_data.get(fieldnames_mapping[condition.replace_where]) or '')

        for encode_regex in self.encode_regex:
            column_value = mapped_data[fieldnames_mapping[encode_regex.replace_where]].strip()
            encoded_value = encode_regex.encode(column_value, encode)
            mapped_data[fieldnames_mapping[encode_regex.replace_where]] = encoded_value

        return mapped_data

    def de_mapper(
        self,
        in_data: dict[str, str],
        worker: Worker,
        _fieldnames_mapping: dict[str, str],
    ) -> dict[str, str]:
        de_encode = worker.encoded_replace

        # Other data types should not be handled, only strings make sense.
        def handler(data: Any) -> Any:
            if isinstance(data, str) or isinstance(data, bytes):
                return ENC_PATTERN.sub(de_encode, data)
            return data

        return {
            key: handler(value)
            for key, value in in_data.items()
        }

    def get_description(self) -> dict[str, str]:
        return self.make_description(
            f'Clear columns: {self.clear_columns or "None"}\nEncode columns: {self.encode_columns or "None"}',
        )


class XlsxReader(csv.DictReader):
    class Reader:
        def __init__(self, worksheet: Worksheet):
            self.generator = worksheet.values
            self.line_num = 0

        def __iter__(self):
            return self

        def __next__(self):
            self.line_num += 1
            return next(self.generator)

    def __init__(self, worksheet: Worksheet, *args, **kwargs):
        super().__init__(f=[], *args, **kwargs)
        self.reader = self.Reader(worksheet)


class XlsxWriter(csv.DictWriter):
    class Writer:
        def __init__(self, original: Worksheet):
            self.workbook = Workbook()
            # Workbook is automatically created with a sheet.
            self.worksheet = self.workbook.active
            self.worksheet.title = original.title
            self.row_index = 1

        def writerow(self, list_of_values: list[Any]) -> int:
            # This is FAR from optimal, but it should work and was easy to write.
            for index, value in enumerate(list_of_values, start=1):
                cell = self.worksheet.cell(row=self.row_index, column=index)
                cell.value = value

            self.row_index += 1
            return 0

        def writerows(self, list_of_list_of_values: list[list[Any]]) -> int:
            for list_of_values in list_of_list_of_values:
                self.writerow(list_of_values)
            return 0

    def __init__(self, worksheet: Worksheet, *args, **kwargs):
        super().__init__(f=io.StringIO(), *args, **kwargs)
        self.writer = self.Writer(worksheet)

    def save_workbook(self, out_stream: BinaryIO) -> None:
        # This is the official way of making a stream out of a workbook.
        # https://openpyxl.readthedocs.io/en/stable/tutorial.html#saving-as-a-stream
        with NamedTemporaryFile() as temp_file:
            self.writer.workbook.save(temp_file.name)
            temp_file.seek(0)
            out_stream.write(temp_file.read())


@ConfigFactory.register
class XLSXConfig(CSVConfig):
    CONFIG_TYPE = 'xlsx-config'
    BUFFER_TYPE = io.BytesIO

    def __init__(
        self,
        clear_columns: Iterable[str],
        encode_columns: Iterable[str],
        encode_conditional: Optional[Iterable[list[str]]] = None,
        encode_regex: Optional[Iterable[list[str]]] = None,
        num_headers: int = 1,
        skip_initial_lines: int = 0,
        **kwargs,
    ):
        super().__init__(
            clear_columns=clear_columns,
            encode_columns=encode_columns,
            encode_conditional=encode_conditional,
            encode_regex=encode_regex,
            dialect='--',
            delimiter=None,
            num_headers=num_headers,
            skip_initial_lines=skip_initial_lines,
            external_header_file=None,
            external_header_format=None,
            **kwargs,
        )

    def make_destination_buffer(self) -> BUFFER_TYPE:
        return io.BytesIO()

    def make_buffer_binary(self, destination_buffer: BUFFER_TYPE) -> bytes:
        return destination_buffer.getvalue()

    @contextlib.contextmanager
    def make_csv_reader_writer(
        self,
        in_file: FilePath,
        destination: io.BytesIO,
    ) -> tuple[csv.DictReader, csv.DictWriter]:
        # When reading an Excel file, it's better to load it all up into the memory.
        # This way we can even load files from inside a zip archive.
        with in_file.open(mode='rb') as f:  # noqa (mode is supported)
            workbook_data = f.read()

        workbook = load_workbook(io.BytesIO(workbook_data), read_only=True,
                                 rich_text=True)  # noqa (rich_text not in pyi)
        worksheet = workbook.active

        reader = XlsxReader(worksheet)
        writer = XlsxWriter(worksheet, fieldnames=reader.fieldnames)

        yield reader, writer

        writer.save_workbook(destination)


@ConfigFactory.register
class RawRegexConfig(BaseConfig):
    CONFIG_TYPE = 'raw-regex-config'
    BUFFER_TYPE = io.TextIOWrapper

    def __init__(self, regex_groups: list[str], **kwargs):
        super().__init__(**kwargs)
        self.regex_groups = [EncodeRegex(expression) for expression in regex_groups]

    def encode_file(self, in_file: FilePath, worker: Worker, destination: BUFFER_TYPE) -> None:
        with in_file.open(mode='r', encoding=self.encoding) as source:  # noqa (encoding is supported)
            for line in source:
                out_line = line
                for regex in self.regex_groups:
                    out_line = regex.encode(out_line, worker.encode_value)
                destination.writelines([out_line])

    def decode_file(self, in_file: FilePath, worker: Worker, destination: BUFFER_TYPE) -> None:
        with in_file.open(mode='r', encoding=self.encoding) as source:  # noqa (encoding is supported)
            content = source.read()
            content = ENC_PATTERN.sub(worker.encoded_replace, content)
            destination.write(content)

    def get_description(self) -> dict[str, str]:
        return self.make_description(
            'Replace each named group in the following expressions:\n\n'
            + '\n'.join([entry.expression for entry in self.regex_groups]),
        )


def add_common_arguments(parser: GooeyParser, add_mapping: bool):
    parser.add_argument(
        'output_directory',
        metavar='Output directory',
        widget='DirChooser',
        help='Path to store output files',
    )
    if add_mapping:
        parser.add_argument(
            'mapping_file',
            metavar='Mapping file',
            widget='FileChooser',
            help='mapping.tsv file',
            gooey_options={
                'wildcard': "Tab separated file (*.tsv)|*.tsv",
            },
        )

    parser.add_argument(
        'input',
        action='store',
        nargs='+',
        metavar='Input files',
        widget='MultiFileChooser',
        help='Files or directories to be processed',
    )


def main():
    parser = GooeyParser(
        description='Program to anonymize data files for Byte Analytics Mobile Optimizer',
        epilog='Run without arguments to launch the GUI',
    )
    encode_tag = 'Encode'
    decode_tag = 'Decode'

    subparsers = parser.add_subparsers(dest='action', required=True)
    encode = subparsers.add_parser(encode_tag, help='Anonymize the data files')
    add_common_arguments(encode, False)

    decode = subparsers.add_parser(decode_tag, help='De-anonymize the data files')
    add_common_arguments(decode, True)

    args = parser.parse_args()

    assert args.action in (encode_tag, decode_tag)
    for_encode = args.action == encode_tag

    with Worker(args.output_directory, should_save_mappings=for_encode) as worker:
        worker.find_files(args.input, for_encode=for_encode)
        if for_encode and (path := Path(args.output_directory) / Worker.MAPPING_FILE_NAME).exists():
            worker.load_mappings(path)
        elif not for_encode:
            worker.load_mappings(args.mapping_file)
        worker.process_files()


def get_resource_path(*args):
    if getattr(sys, 'frozen', False):
        # MEIPASS explanation:
        # https://pythonhosted.org/PyInstaller/#run-time-operation
        resource_dir = getattr(sys, '_MEIPASS', None)
    else:
        resource_dir = os.path.normpath(os.path.dirname(__file__))
    return os.path.join(resource_dir, *args)


if __name__ == '__main__':
    if len(sys.argv) > 1:
        # CLI
        IGNORE_COMMAND = '--ignore-gooey'
        if IGNORE_COMMAND in sys.argv:
            sys.argv.remove(IGNORE_COMMAND)
        main()
    else:
        MENU = [
            {'name': 'Help', 'items': [
                {
                    'type': 'AboutDialog',
                    'menuTitle': 'About',
                    'name': 'Byte Analytics Data Anonymizer',
                    'description': 'Program to anonymize data files for Byte Analytics Mobile Optimizer',
                    'version': 'latest',  # TODO: git tag
                    'copyright': '2021',
                    'website': 'https://github.com/reef-technologies/byteanalytics-anonymizer',
                    'developer': 'https://reef.pl/',
                    'license': 'GPL v3'
                },
                {
                    'type': 'Link',
                    'menuTitle': 'Check for updates',
                    'url': 'https://github.com/reef-technologies/byteanalytics-anonymizer/releases',
                }
            ]},
            {'name': 'Carrier Configuration', 'items': list(ConfigFactory.get_config_descriptions())},
        ]

        # GUI
        Gooey(
            f=main,
            language='english',
            show_sidebar=True,
            program_name='Byte Analytics Data Encoder',
            advanced=True,
            default_size=(900, 600),
            required_cols=1,
            optional_cols=1,
            menu=MENU,
            image_dir=get_resource_path('images'),
            language_dir=get_resource_path('gooey', 'languages'),
            progress_regex=r"^Progress (?P<percent>\d+)%$",
            progress_expr="percent",
            disable_progress_bar_animation=True,
            hide_progress_msg=True,
        )()
