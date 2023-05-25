import io
import pathlib

import pytest
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from anonymizer import XlsxReader


@pytest.fixture
def xlsx_file() -> pathlib.Path:
    return pathlib.Path(__file__).parent / 'data/bell/test-Hardware report.xlsx'


def assert_valid_workbook(workbook: Workbook) -> None:
    worksheet = workbook.active

    reader = XlsxReader(worksheet)
    lines = []
    for dict_line in reader:
        lines.append(dict_line)

    assert len(lines) == 3

    # Check some basic fields that we know from this document.
    field_tag_mapping = [
        ('Mobile number', 'test-mobile-{}'),
        ('User last name', 'test-last-{}'),
        ('User first name', 'test-first-{}'),
        ('Model code', 'test-model-{}'),
    ]

    for idx, line_dict in enumerate(lines, start=1):
        for dict_key, format_tag in field_tag_mapping:
            assert line_dict.get(dict_key) == format_tag.format(idx)


def test_read_xlsx_file(xlsx_file) -> None:
    workbook = load_workbook(xlsx_file, read_only=True, rich_text=True)  # noqa: rich_text missing from pyi
    assert_valid_workbook(workbook)


def test_read_xlsx_from_bytes(xlsx_file) -> None:
    file_data = xlsx_file.read_bytes()
    buffer = io.BytesIO(file_data)
    workbook = load_workbook(buffer, read_only=True, rich_text=True)  # noqa: rich_text missing from pyi
    assert_valid_workbook(workbook)
