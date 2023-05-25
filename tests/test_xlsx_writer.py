import io
import pathlib

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from anonymizer import XlsxReader, XlsxWriter


def test_xlsx_writer(fake_fs):
    input_workbook = Workbook()
    input_worksheet = input_workbook.create_sheet(title='test-title')

    header = ['test1', 'test2', 'test3']
    data = [
        ['value1', 'value2', 'value3'],
        ['value4', 'value5', 'value6'],
        ['value7', 'value8', 'value9'],
    ]

    writer = XlsxWriter(input_worksheet, fieldnames=['test1', 'test2', 'test3'])
    writer.writeheader()
    for data_line in data:
        mapping = dict(zip(header, data_line))
        writer.writerow(mapping)

    output_file = 'test.xlsx'
    with open(output_file, 'wb') as f:
        writer.save_workbook(f)

    def assert_workbook(workbook: Workbook) -> None:
        worksheet = workbook.active
        assert worksheet.title == input_worksheet.title

        # At this point it is assumed that reader is "ok".
        reader = XlsxReader(worksheet=worksheet)
        for idx, line_dict in enumerate(reader):
            expected_line = dict(zip(header, data[idx]))
            assert line_dict == expected_line

    # Check loading from file.
    file_workbook = load_workbook(output_file, read_only=True, rich_text=True)  # noqa: rich_text missing from pyi
    assert_workbook(file_workbook)

    # Also check loading only from a binary representation of the file.
    buffer = pathlib.Path(output_file).read_bytes()
    binary_workbook = load_workbook(io.BytesIO(buffer), read_only=True, rich_text=True)  # noqa: rich_text missing from pyi
    assert_workbook(binary_workbook)
